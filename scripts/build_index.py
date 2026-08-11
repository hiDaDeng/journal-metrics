#!/usr/bin/env python3
"""Build a fast SQLite index from the two source workbooks."""

from __future__ import annotations

import argparse
import json
import os
import re
import sqlite3
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_name(value: object) -> str:
    text = unicodedata.normalize("NFKC", clean(value)).casefold()
    return "".join(char for char in text if char.isalnum())


def normalize_issn(value: object) -> str:
    return re.sub(r"[^0-9X]", "", clean(value).upper())


def header_map(row: tuple[object, ...]) -> dict[str, int]:
    return {clean(value): index for index, value in enumerate(row) if clean(value)}


def get(row: tuple[object, ...], headers: dict[str, int], name: str) -> str:
    index = headers.get(name)
    return clean(row[index]) if index is not None and index < len(row) else ""


def create_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        CREATE TABLE jcr_metrics (
            year INTEGER NOT NULL,
            journal TEXT NOT NULL,
            journal_norm TEXT NOT NULL,
            jif TEXT,
            jcr_quartile TEXT,
            issn TEXT,
            issn_norm TEXT,
            eissn TEXT,
            eissn_norm TEXT,
            category TEXT,
            source_sheet TEXT NOT NULL,
            source_row INTEGER NOT NULL
        );

        CREATE TABLE emerging_quartiles (
            year INTEGER NOT NULL,
            journal TEXT NOT NULL,
            journal_norm TEXT NOT NULL,
            issn1 TEXT,
            issn1_norm TEXT,
            issn2 TEXT,
            issn2_norm TEXT,
            major_category TEXT,
            emerging_quartile TEXT,
            source_row INTEGER NOT NULL
        );

        CREATE TABLE metadata (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );

        CREATE INDEX idx_jcr_name ON jcr_metrics(journal_norm);
        CREATE INDEX idx_jcr_issn ON jcr_metrics(issn_norm);
        CREATE INDEX idx_jcr_eissn ON jcr_metrics(eissn_norm);
        CREATE INDEX idx_emerging_name ON emerging_quartiles(journal_norm);
        CREATE INDEX idx_emerging_issn1 ON emerging_quartiles(issn1_norm);
        CREATE INDEX idx_emerging_issn2 ON emerging_quartiles(issn2_norm);
        """
    )


def load_jcr(connection: sqlite3.Connection, path: Path) -> tuple[int, list[int]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    inserted = 0
    years: list[int] = []
    statement = """
        INSERT INTO jcr_metrics (
            year, journal, journal_norm, jif, jcr_quartile,
            issn, issn_norm, eissn, eissn_norm, category,
            source_sheet, source_row
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """

    for sheet_name in workbook.sheetnames:
        try:
            year = int(clean(sheet_name))
        except ValueError:
            continue
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = header_map(next(rows))
        except StopIteration:
            continue
        required = {"期刊", "JIF", "JCR分区", "ISSN", "eISSN", "所属类别"}
        missing = required.difference(headers)
        if missing:
            raise ValueError(f"{path.name}/{sheet_name} 缺少字段: {sorted(missing)}")

        batch = []
        for row_number, row in enumerate(rows, start=2):
            journal = get(row, headers, "期刊")
            if not journal:
                continue
            issn = get(row, headers, "ISSN")
            eissn = get(row, headers, "eISSN")
            batch.append(
                (
                    year,
                    journal,
                    normalize_name(journal),
                    get(row, headers, "JIF"),
                    get(row, headers, "JCR分区"),
                    issn,
                    normalize_issn(issn),
                    eissn,
                    normalize_issn(eissn),
                    get(row, headers, "所属类别"),
                    sheet_name,
                    row_number,
                )
            )
            if len(batch) >= 2000:
                connection.executemany(statement, batch)
                inserted += len(batch)
                batch.clear()
        if batch:
            connection.executemany(statement, batch)
            inserted += len(batch)
        years.append(year)

    workbook.close()
    return inserted, sorted(set(years))


def load_emerging(connection: sqlite3.Connection, path: Path) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = header_map(next(rows))
    except StopIteration as error:
        raise ValueError(f"{path.name} 是空工作簿") from error
    required = {"大类", "期刊名称", "issn1", "issn2", "分区"}
    missing = required.difference(headers)
    if missing:
        raise ValueError(f"{path.name} 缺少字段: {sorted(missing)}")

    statement = """
        INSERT INTO emerging_quartiles (
            year, journal, journal_norm, issn1, issn1_norm,
            issn2, issn2_norm, major_category, emerging_quartile, source_row
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """
    batch = []
    inserted = 0
    for row_number, row in enumerate(rows, start=2):
        journal = get(row, headers, "期刊名称")
        if not journal:
            continue
        issn1 = get(row, headers, "issn1")
        issn2 = get(row, headers, "issn2")
        batch.append(
            (
                2026,
                journal,
                normalize_name(journal),
                issn1,
                normalize_issn(issn1),
                issn2,
                normalize_issn(issn2),
                get(row, headers, "大类"),
                get(row, headers, "分区"),
                row_number,
            )
        )
        if len(batch) >= 2000:
            connection.executemany(statement, batch)
            inserted += len(batch)
            batch.clear()
    if batch:
        connection.executemany(statement, batch)
        inserted += len(batch)
    workbook.close()
    return inserted


def build(jcr_path: Path, emerging_path: Path, output: Path) -> dict[str, object]:
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_suffix(output.suffix + ".tmp")
    if temporary.exists():
        temporary.unlink()

    connection = sqlite3.connect(temporary)
    try:
        connection.execute("PRAGMA journal_mode = OFF")
        connection.execute("PRAGMA synchronous = OFF")
        create_schema(connection)
        jcr_rows, years = load_jcr(connection, jcr_path)
        emerging_rows = load_emerging(connection, emerging_path)
        metadata = {
            "built_at_utc": datetime.now(timezone.utc).isoformat(),
            "jcr_source": jcr_path.name,
            "emerging_source": emerging_path.name,
            "jcr_years": json.dumps(years),
            "latest_three_jcr_years": json.dumps(years[-3:]),
            "emerging_year": "2026",
        }
        connection.executemany(
            "INSERT INTO metadata(key, value) VALUES (?, ?)", metadata.items()
        )
        connection.commit()
        integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
        if integrity != "ok":
            raise RuntimeError(f"SQLite integrity check failed: {integrity}")
    finally:
        connection.close()

    os.replace(temporary, output)
    return {
        "output": str(output),
        "jcr_rows": jcr_rows,
        "emerging_rows": emerging_rows,
        "jcr_years": years,
        "latest_three_jcr_years": years[-3:],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("jcr_workbook", type=Path)
    parser.add_argument("emerging_workbook", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    result = build(args.jcr_workbook, args.emerging_workbook, args.output)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
